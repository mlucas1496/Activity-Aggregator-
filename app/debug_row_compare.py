"""Compare mapped and not-mapped rows between our output and the target."""
import openpyxl
import os, glob
from collections import Counter

TARGET = "/Users/matthew.lucas/Downloads/Activity Aggregator - 03.06.2026 Forecast .xlsx"
OURS_DIR = "/Users/matthew.lucas/Documents/Activity Aggregator Update/app/output/"

outputs = glob.glob(os.path.join(OURS_DIR, "*.xlsx"))
OURS = max(outputs, key=os.path.getmtime)
print(f"Our output: {OURS}")
print(f"Target:     {TARGET}")

def load_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Alteryx_Output"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i for i, h in enumerate(headers) if h}

    rows_by_status = {"Mapped": [], "Not Mapped - Check": [], "Exclude - Pass": []}
    for row in ws.iter_rows(min_row=2, values_only=True):
        muc = row[col_idx.get("Manual User Check", -1)] or ""
        if muc in rows_by_status:
            rows_by_status[muc].append(row)
    wb.close()
    return headers, col_idx, rows_by_status

print("\nLoading target...")
t_h, t_idx, t_rows = load_rows(TARGET)
print(f"  Mapped: {len(t_rows['Mapped'])}, NotMapped: {len(t_rows['Not Mapped - Check'])}, Excluded: {len(t_rows['Exclude - Pass'])}")

print("Loading ours...")
o_h, o_idx, o_rows = load_rows(OURS)
print(f"  Mapped: {len(o_rows['Mapped'])}, NotMapped: {len(o_rows['Not Mapped - Check'])}, Excluded: {len(o_rows['Exclude - Pass'])}")

def make_key(row, idx):
    """Create a comparison key from core identifying fields."""
    source = str(row[idx.get("Source", -1)] or "").strip()
    acct = str(row[idx.get("Account Name", -1)] or "").strip()
    asset = str(row[idx.get("Asset Code", -1)] or "").strip()
    net = row[idx.get("Net Activity - USD", -1)]
    try:
        net = round(float(net), 2)
    except:
        net = 0
    notes = str(row[idx.get("Notes", -1)] or "")[:80]
    ref = str(row[idx.get("13WCF Ref #", -1)] or "").strip()
    return (source, acct, asset, net, notes, ref)

def make_key_with_mapping(row, idx):
    """Key including the mapping for category comparison."""
    base = make_key(row, idx)
    mapping = str(row[idx.get("13WCF Line Item Mapping", -1)] or "").strip()
    return base + (mapping,)

# ============ MAPPED ROWS ============
print("\n" + "="*80)
print("MAPPED ROWS COMPARISON")
print("="*80)

t_mapped_keys = Counter(make_key(r, t_idx) for r in t_rows["Mapped"])
o_mapped_keys = Counter(make_key(r, o_idx) for r in o_rows["Mapped"])

# Rows in ours but not target (extra mapped)
extra_mapped = o_mapped_keys - t_mapped_keys
print(f"\nRows MAPPED in ours but NOT in target ({sum(extra_mapped.values())}):")
for key, count in sorted(extra_mapped.items()):
    print(f"  [{count}x] Source={key[0]}, Acct={key[1]}, Asset={key[2]}, Net={key[3]}, Notes={key[4][:50]}, Ref={key[5]}")

# Rows in target but not ours (missing mapped)
missing_mapped = t_mapped_keys - o_mapped_keys
print(f"\nRows MAPPED in target but NOT in ours ({sum(missing_mapped.values())}):")
for key, count in sorted(missing_mapped.items()):
    print(f"  [{count}x] Source={key[0]}, Acct={key[1]}, Asset={key[2]}, Net={key[3]}, Notes={key[4][:50]}, Ref={key[5]}")

# Check category mismatches (rows mapped in both but different mapping)
print(f"\nCategory mismatches (mapped in both but different 13WCF Line Item Mapping):")
t_map_with_cat = {}
for r in t_rows["Mapped"]:
    k = make_key(r, t_idx)
    m = str(r[t_idx.get("13WCF Line Item Mapping", -1)] or "").strip()
    t_map_with_cat[k] = m

o_map_with_cat = {}
for r in o_rows["Mapped"]:
    k = make_key(r, o_idx)
    m = str(r[o_idx.get("13WCF Line Item Mapping", -1)] or "").strip()
    o_map_with_cat[k] = m

mismatch_count = 0
for k in set(t_map_with_cat.keys()) & set(o_map_with_cat.keys()):
    if t_map_with_cat[k] != o_map_with_cat[k]:
        mismatch_count += 1
        print(f"  Acct={k[1][:40]}, Asset={k[2]}, Net={k[3]}")
        print(f"    Target: {t_map_with_cat[k]}")
        print(f"    Ours:   {o_map_with_cat[k]}")
if mismatch_count == 0:
    print("  None - all shared mapped rows have same category!")

# ============ NOT MAPPED ROWS ============
print("\n" + "="*80)
print("NOT MAPPED ROWS COMPARISON")
print("="*80)

t_nm_keys = Counter(make_key(r, t_idx) for r in t_rows["Not Mapped - Check"])
o_nm_keys = Counter(make_key(r, o_idx) for r in o_rows["Not Mapped - Check"])

extra_nm = o_nm_keys - t_nm_keys
missing_nm = t_nm_keys - o_nm_keys

print(f"\nNot-mapped in ours but NOT in target: {sum(extra_nm.values())}")
# Show first 20
for i, (key, count) in enumerate(sorted(extra_nm.items())):
    if i >= 20:
        print(f"  ... and {sum(extra_nm.values()) - 20} more")
        break
    print(f"  [{count}x] Source={key[0]}, Acct={key[1][:40]}, Asset={key[2]}, Net={key[3]}, Ref={key[5]}")

print(f"\nNot-mapped in target but NOT in ours: {sum(missing_nm.values())}")
for i, (key, count) in enumerate(sorted(missing_nm.items())):
    if i >= 20:
        print(f"  ... and {sum(missing_nm.values()) - 20} more")
        break
    print(f"  [{count}x] Source={key[0]}, Acct={key[1][:40]}, Asset={key[2]}, Net={key[3]}, Ref={key[5]}")

# ============ EXCLUDED ROWS ============
print("\n" + "="*80)
print("EXCLUDED ROWS COMPARISON")
print("="*80)
print(f"\nTarget excluded: {len(t_rows['Exclude - Pass'])}")
print(f"Ours excluded:   {len(o_rows['Exclude - Pass'])}")
print(f"Difference:      {len(o_rows['Exclude - Pass']) - len(t_rows['Exclude - Pass'])}")

# Check if the extra excluded are all Bitgo_2025_v2
extra_excl_accts = Counter()
t_excl_keys = Counter(make_key(r, t_idx) for r in t_rows["Exclude - Pass"])
o_excl_keys = Counter(make_key(r, o_idx) for r in o_rows["Exclude - Pass"])
extra_excl = o_excl_keys - t_excl_keys
for key, count in extra_excl.items():
    extra_excl_accts[key[1]] += count

print(f"\nExtra excluded rows by Account Name:")
for acct, count in extra_excl_accts.most_common(10):
    print(f"  {acct}: {count}")

print(f"\nTotal extra excluded: {sum(extra_excl_accts.values())}")
