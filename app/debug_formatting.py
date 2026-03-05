#!/usr/bin/env python3
"""
debug_formatting.py
Extracts all formatting details from the target Excel file so we can replicate
the exact look in our generated output.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Alignment, numbers

TARGET = "/Users/matthew.lucas/Downloads/Activity Aggregator - 03.06.2026 Forecast .xlsx"

SEP = "=" * 90
SUBSEP = "-" * 70


def color_str(color):
    """Return a human-readable representation of a Color object."""
    if color is None:
        return "None"
    if color.type == "rgb" and color.rgb:
        return f"rgb={color.rgb}"
    if color.type == "theme":
        return f"theme={color.theme}, tint={color.tint}"
    if color.type == "indexed":
        return f"indexed={color.indexed}"
    return str(color)


def font_str(f):
    if f is None:
        return "None"
    parts = []
    parts.append(f"name={f.name!r}")
    parts.append(f"sz={f.sz}")
    parts.append(f"bold={f.bold}")
    parts.append(f"italic={f.italic}")
    parts.append(f"underline={f.underline}")
    parts.append(f"strike={f.strike}")
    parts.append(f"color={color_str(f.color)}")
    return ", ".join(parts)


def fill_str(f):
    if f is None:
        return "None"
    parts = []
    parts.append(f"patternType={f.patternType!r}")
    parts.append(f"fgColor={color_str(f.fgColor)}")
    parts.append(f"bgColor={color_str(f.bgColor)}")
    return ", ".join(parts)


def border_side_str(side):
    if side is None:
        return "None"
    return f"style={side.style!r}, color={color_str(side.color)}"


def border_str(b):
    if b is None:
        return "None"
    parts = []
    parts.append(f"  left:   {border_side_str(b.left)}")
    parts.append(f"  right:  {border_side_str(b.right)}")
    parts.append(f"  top:    {border_side_str(b.top)}")
    parts.append(f"  bottom: {border_side_str(b.bottom)}")
    return "\n".join(parts)


def alignment_str(a):
    if a is None:
        return "None"
    parts = []
    parts.append(f"horizontal={a.horizontal}")
    parts.append(f"vertical={a.vertical}")
    parts.append(f"wrapText={a.wrapText}")
    parts.append(f"shrinkToFit={a.shrinkToFit}")
    parts.append(f"textRotation={a.textRotation}")
    parts.append(f"indent={a.indent}")
    return ", ".join(parts)


def print_cell_format(cell, label=""):
    prefix = f"  [{label}] " if label else "  "
    val_repr = repr(cell.value)
    if len(val_repr) > 60:
        val_repr = val_repr[:60] + "..."
    print(f"{prefix}Value: {val_repr}")
    print(f"{prefix}NumberFormat: {cell.number_format!r}")
    print(f"{prefix}Font: {font_str(cell.font)}")
    print(f"{prefix}Fill: {fill_str(cell.fill)}")
    print(f"{prefix}Alignment: {alignment_str(cell.alignment)}")
    print(f"{prefix}Border:\n{prefix}" + border_str(cell.border).replace("\n", f"\n{prefix}"))


# ---------------------------------------------------------------------------
# Load workbook
# ---------------------------------------------------------------------------
print(f"Loading: {TARGET}")
wb = openpyxl.load_workbook(TARGET, data_only=False)
print(f"Sheet names: {wb.sheetnames}\n")


# ###########################################################################
# 1. Alteryx_Output
# ###########################################################################
print(SEP)
print("SHEET: Alteryx_Output")
print(SEP)

ws = wb["Alteryx_Output"]

# --- dimensions ---
print(f"\nDimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# --- Freeze panes ---
print(f"\nFreeze panes: {ws.freeze_panes}")

# --- Auto filter ---
print(f"Auto filter ref: {ws.auto_filter.ref}")

# --- Sheet view / zoom ---
for i, sv in enumerate(ws.views.sheetView):
    print(f"SheetView[{i}]: tabSelected={sv.tabSelected}, zoomScale={sv.zoomScale}, "
          f"showGridLines={sv.showGridLines}, defaultGridColor={sv.defaultGridColor}")

# --- Column widths ---
print(f"\n{SUBSEP}")
print("Column widths (Alteryx_Output)")
print(SUBSEP)
for col_idx in range(1, ws.max_column + 1):
    letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    if dim:
        print(f"  Col {col_idx:>2} ({letter:>2}): width={dim.width}, "
              f"bestFit={dim.bestFit}, hidden={dim.hidden}, "
              f"customWidth={dim.customWidth}")
    else:
        print(f"  Col {col_idx:>2} ({letter:>2}): <default>")

# --- Row heights ---
print(f"\n{SUBSEP}")
print("Row heights (first 10 rows)")
print(SUBSEP)
for r in range(1, min(11, ws.max_row + 1)):
    dim = ws.row_dimensions.get(r)
    if dim:
        print(f"  Row {r}: height={dim.height}, customHeight={dim.customHeight}, hidden={dim.hidden}")
    else:
        print(f"  Row {r}: <default>")

# --- Header row (row 1) ---
print(f"\n{SUBSEP}")
print("Header row formatting (row 1)")
print(SUBSEP)
for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col_idx)
    letter = get_column_letter(col_idx)
    print(f"\n  Column {col_idx} ({letter}): header = {cell.value!r}")
    print_cell_format(cell, f"Col{col_idx}")

# --- Data rows (rows 2-6) ---
print(f"\n{SUBSEP}")
print("Data row formatting (rows 2-6, first 5 data rows)")
print(SUBSEP)
for row_idx in range(2, min(7, ws.max_row + 1)):
    print(f"\n  --- Row {row_idx} ---")
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        letter = get_column_letter(col_idx)
        # Only print full detail if something interesting
        print(f"\n    Col {col_idx} ({letter}):")
        val_repr = repr(cell.value)
        if len(val_repr) > 60:
            val_repr = val_repr[:60] + "..."
        print(f"      Value: {val_repr}")
        print(f"      NumFmt: {cell.number_format!r}")
        print(f"      Font: {font_str(cell.font)}")
        print(f"      Fill: {fill_str(cell.fill)}")
        print(f"      Alignment: {alignment_str(cell.alignment)}")

# --- Conditional formatting ---
print(f"\n{SUBSEP}")
print("Conditional formatting rules (Alteryx_Output)")
print(SUBSEP)
if ws.conditional_formatting:
    for cf in ws.conditional_formatting:
        print(f"\n  Range(s): {cf}")
        for rule in cf.rules:
            print(f"    Type: {rule.type}, Priority: {rule.priority}")
            print(f"    Formula: {rule.formula}")
            print(f"    Operator: {rule.operator}")
            if rule.dxf:
                if rule.dxf.font:
                    print(f"    DXF Font: {font_str(rule.dxf.font)}")
                if rule.dxf.fill:
                    print(f"    DXF Fill: {fill_str(rule.dxf.fill)}")
                if rule.dxf.border:
                    print(f"    DXF Border present: True")
                if rule.dxf.numFmt:
                    print(f"    DXF NumFmt: {rule.dxf.numFmt}")
else:
    print("  (none)")

# --- Merged cells ---
print(f"\n  Merged cells: {list(ws.merged_cells.ranges)}")


# ###########################################################################
# 2. WD - Find Bank Statement Lines
# ###########################################################################
print(f"\n\n{SEP}")
print("SHEET: WD - Find Bank Statement Lines")
print(SEP)

ws2 = wb["WD - Find Bank Statement Lines"]

print(f"\nDimensions: {ws2.dimensions}")
print(f"Max row: {ws2.max_row}, Max col: {ws2.max_column}")
print(f"Freeze panes: {ws2.freeze_panes}")
print(f"Auto filter ref: {ws2.auto_filter.ref}")

for i, sv in enumerate(ws2.views.sheetView):
    print(f"SheetView[{i}]: tabSelected={sv.tabSelected}, zoomScale={sv.zoomScale}, "
          f"showGridLines={sv.showGridLines}")

# --- Column widths ---
print(f"\n{SUBSEP}")
print("Column widths (WD - Find Bank Statement Lines)")
print(SUBSEP)
for col_idx in range(1, ws2.max_column + 1):
    letter = get_column_letter(col_idx)
    dim = ws2.column_dimensions.get(letter)
    if dim:
        print(f"  Col {col_idx:>2} ({letter:>2}): width={dim.width}, "
              f"bestFit={dim.bestFit}, hidden={dim.hidden}, "
              f"customWidth={dim.customWidth}")
    else:
        print(f"  Col {col_idx:>2} ({letter:>2}): <default>")

# --- Row heights (rows 1-15) ---
print(f"\n{SUBSEP}")
print("Row heights (rows 1-15)")
print(SUBSEP)
for r in range(1, 16):
    dim = ws2.row_dimensions.get(r)
    if dim:
        print(f"  Row {r}: height={dim.height}, customHeight={dim.customHeight}, hidden={dim.hidden}")
    else:
        print(f"  Row {r}: <default>")

# --- Merged cells ---
print(f"\n  Merged cells: {list(ws2.merged_cells.ranges)}")

# --- Row 3 (title) ---
print(f"\n{SUBSEP}")
print("Row 3 (title row) formatting")
print(SUBSEP)
for col_idx in range(1, min(ws2.max_column + 1, 30)):
    cell = ws2.cell(row=3, column=col_idx)
    if cell.value is not None:
        print(f"\n  Col {col_idx} ({get_column_letter(col_idx)}):")
        print_cell_format(cell)

# --- Row 6 (Workday / A&M labels) ---
print(f"\n{SUBSEP}")
print("Row 6 (Workday / A&M labels) formatting")
print(SUBSEP)
for col_idx in range(1, min(ws2.max_column + 1, 30)):
    cell = ws2.cell(row=6, column=col_idx)
    if cell.value is not None:
        print(f"\n  Col {col_idx} ({get_column_letter(col_idx)}):")
        print_cell_format(cell)

# --- Rows 7-8 (headers) ---
print(f"\n{SUBSEP}")
print("Rows 7-8 (header rows) formatting")
print(SUBSEP)
for row_idx in [7, 8]:
    print(f"\n  === Row {row_idx} ===")
    for col_idx in range(1, min(ws2.max_column + 1, 30)):
        cell = ws2.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            print(f"\n  Col {col_idx} ({get_column_letter(col_idx)}):")
            print_cell_format(cell)

# --- Data rows (first 5 data rows after headers) ---
print(f"\n{SUBSEP}")
print("Data row formatting (rows 9-13)")
print(SUBSEP)
for row_idx in range(9, min(14, ws2.max_row + 1)):
    print(f"\n  --- Row {row_idx} ---")
    for col_idx in range(1, min(ws2.max_column + 1, 30)):
        cell = ws2.cell(row=row_idx, column=col_idx)
        if cell.value is not None or col_idx <= 15:
            letter = get_column_letter(col_idx)
            val_repr = repr(cell.value)
            if len(val_repr) > 50:
                val_repr = val_repr[:50] + "..."
            print(f"    Col {col_idx} ({letter}): val={val_repr}, "
                  f"numFmt={cell.number_format!r}, "
                  f"font=({font_str(cell.font)}), "
                  f"fill=({fill_str(cell.fill)}), "
                  f"align=({alignment_str(cell.alignment)})")

# --- Conditional formatting ---
print(f"\n{SUBSEP}")
print("Conditional formatting (WD - Find Bank Statement Lines)")
print(SUBSEP)
if ws2.conditional_formatting:
    for cf in ws2.conditional_formatting:
        print(f"\n  Range(s): {cf}")
        for rule in cf.rules:
            print(f"    Type: {rule.type}, Priority: {rule.priority}")
            print(f"    Formula: {rule.formula}")
            print(f"    Operator: {rule.operator}")
            if rule.dxf:
                if rule.dxf.font:
                    print(f"    DXF Font: {font_str(rule.dxf.font)}")
                if rule.dxf.fill:
                    print(f"    DXF Fill: {fill_str(rule.dxf.fill)}")
else:
    print("  (none)")


# ###########################################################################
# 3. BTH Trade Log
# ###########################################################################
print(f"\n\n{SEP}")
print("SHEET: BTH Trade Log")
print(SEP)

ws3 = wb["BTH Trade Log"]

print(f"\nDimensions: {ws3.dimensions}")
print(f"Max row: {ws3.max_row}, Max col: {ws3.max_column}")
print(f"Freeze panes: {ws3.freeze_panes}")
print(f"Auto filter ref: {ws3.auto_filter.ref}")

for i, sv in enumerate(ws3.views.sheetView):
    print(f"SheetView[{i}]: tabSelected={sv.tabSelected}, zoomScale={sv.zoomScale}, "
          f"showGridLines={sv.showGridLines}")

# --- Column widths ---
print(f"\n{SUBSEP}")
print("Column widths (BTH Trade Log)")
print(SUBSEP)
for col_idx in range(1, ws3.max_column + 1):
    letter = get_column_letter(col_idx)
    dim = ws3.column_dimensions.get(letter)
    if dim:
        print(f"  Col {col_idx:>2} ({letter:>2}): width={dim.width}, "
              f"bestFit={dim.bestFit}, hidden={dim.hidden}, "
              f"customWidth={dim.customWidth}")
    else:
        print(f"  Col {col_idx:>2} ({letter:>2}): <default>")

# --- Row heights ---
print(f"\n{SUBSEP}")
print("Row heights (rows 1-15)")
print(SUBSEP)
for r in range(1, 16):
    dim = ws3.row_dimensions.get(r)
    if dim:
        print(f"  Row {r}: height={dim.height}, customHeight={dim.customHeight}, hidden={dim.hidden}")
    else:
        print(f"  Row {r}: <default>")

# --- Merged cells ---
print(f"\n  Merged cells: {list(ws3.merged_cells.ranges)}")

# --- Rows 1-7 (pre-header area) ---
print(f"\n{SUBSEP}")
print("Pre-header rows (1-7) - checking for content")
print(SUBSEP)
for row_idx in range(1, 8):
    for col_idx in range(1, min(ws3.max_column + 1, 20)):
        cell = ws3.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            print(f"  Row {row_idx}, Col {col_idx} ({get_column_letter(col_idx)}):")
            print_cell_format(cell)

# --- Rows 8-9 (headers) ---
print(f"\n{SUBSEP}")
print("Rows 8-9 (header rows) formatting")
print(SUBSEP)
for row_idx in [8, 9]:
    print(f"\n  === Row {row_idx} ===")
    for col_idx in range(1, ws3.max_column + 1):
        cell = ws3.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            print(f"\n  Col {col_idx} ({get_column_letter(col_idx)}):")
            print_cell_format(cell)

# --- Data rows (first 5 after headers) ---
print(f"\n{SUBSEP}")
print("Data row formatting (rows 10-14)")
print(SUBSEP)
for row_idx in range(10, min(15, ws3.max_row + 1)):
    print(f"\n  --- Row {row_idx} ---")
    for col_idx in range(1, ws3.max_column + 1):
        cell = ws3.cell(row=row_idx, column=col_idx)
        if cell.value is not None or col_idx <= 10:
            letter = get_column_letter(col_idx)
            val_repr = repr(cell.value)
            if len(val_repr) > 50:
                val_repr = val_repr[:50] + "..."
            print(f"    Col {col_idx} ({letter}): val={val_repr}, "
                  f"numFmt={cell.number_format!r}, "
                  f"font=({font_str(cell.font)}), "
                  f"fill=({fill_str(cell.fill)}), "
                  f"align=({alignment_str(cell.alignment)})")

# --- Conditional formatting ---
print(f"\n{SUBSEP}")
print("Conditional formatting (BTH Trade Log)")
print(SUBSEP)
if ws3.conditional_formatting:
    for cf in ws3.conditional_formatting:
        print(f"\n  Range(s): {cf}")
        for rule in cf.rules:
            print(f"    Type: {rule.type}, Priority: {rule.priority}")
            print(f"    Formula: {rule.formula}")
            print(f"    Operator: {rule.operator}")
            if rule.dxf:
                if rule.dxf.font:
                    print(f"    DXF Font: {font_str(rule.dxf.font)}")
                if rule.dxf.fill:
                    print(f"    DXF Fill: {fill_str(rule.dxf.fill)}")
else:
    print("  (none)")


# ###########################################################################
# Summary of special number formats across ALL sheets
# ###########################################################################
print(f"\n\n{SEP}")
print("SUMMARY: Distinct number formats per sheet")
print(SEP)

for sname in ["Alteryx_Output", "WD - Find Bank Statement Lines", "BTH Trade Log"]:
    ws_tmp = wb[sname]
    fmts = {}
    for row in ws_tmp.iter_rows(min_row=1, max_row=min(ws_tmp.max_row, 500)):
        for cell in row:
            nf = cell.number_format
            if nf not in fmts:
                fmts[nf] = []
            if len(fmts[nf]) < 3:
                fmts[nf].append(f"{get_column_letter(cell.column)}{cell.row}")
    print(f"\n  Sheet: {sname}")
    for nf, examples in sorted(fmts.items()):
        print(f"    {nf!r:40s}  examples: {', '.join(examples)}")

print(f"\n\n{SEP}")
print("DONE - formatting extraction complete")
print(SEP)
