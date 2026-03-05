"""
Debug script: Analyze the target Alteryx output to understand the 431 vs 429
mapped-row discrepancy after the intercompany detection change in stage4_lukka.py.

The change: intercompany detection now only sums Net Activity - USD for UNMAPPED
rows when checking if a 13WCF Ref # group sums to ~0. This could cause ref groups
that previously did NOT sum to ~0 (because mapped rows with large values were
included) to now sum to ~0, tagging extra rows as Intercompany Inflow/Outflow.

This script examines the TARGET (Alteryx) output to see:
  1. Total mapped rows and their distribution by Source / 13WCF Line Item Mapping
  2. Specifically how many Intercompany Inflow/Outflow rows exist in the target
  3. Which 13WCF Ref # groups contain intercompany rows, with their Net Activity sums
"""

import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Configuration ──
TARGET_FILE = "/Users/matthew.lucas/Downloads/Activity Aggregator - 03.06.2026 Forecast .xlsx"
SHEET_NAME = "Alteryx_Output"


def main():
    print(f"Opening: {TARGET_FILE}")
    print(f"Sheet:   {SHEET_NAME}")
    print("=" * 80)

    wb = openpyxl.load_workbook(TARGET_FILE, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        wb.close()
        sys.exit(1)

    ws = wb[SHEET_NAME]

    # Read header row (row 1)
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter)
    headers = [str(h).strip() if h is not None else "" for h in headers]

    # Build column index map
    col_idx = {name: i for i, name in enumerate(headers)}

    # Key columns we need
    KEY_COLS = [
        "Source", "Account Name", "Asset Code", "Notes", "Notes_Clean",
        "13WCF Line Item Mapping", "13WCF Ref #", "Net Activity - USD",
        "Manual User Check", "Incl/Excl", "R/D", "Type", "Sub Type",
        "Counter_Asset_Exists", "Counter Asset Base Code", "Matched_Substring",
    ]

    missing_cols = [c for c in KEY_COLS if c not in col_idx]
    if missing_cols:
        print(f"WARNING: Missing columns: {missing_cols}")

    # Read all data rows into list of dicts
    all_data = []
    for row_vals in rows_iter:
        record = {}
        for col_name in KEY_COLS:
            if col_name in col_idx:
                val = row_vals[col_idx[col_name]] if col_idx[col_name] < len(row_vals) else None
                record[col_name] = val
        all_data.append(record)

    wb.close()

    print(f"Total data rows in Alteryx_Output: {len(all_data)}")
    print()

    # ── 1. Count by Manual User Check ──
    check_counts = defaultdict(int)
    for row in all_data:
        status = str(row.get("Manual User Check") or "").strip()
        check_counts[status] += 1

    print("1) MANUAL USER CHECK DISTRIBUTION")
    print("-" * 50)
    for status, count in sorted(check_counts.items(), key=lambda x: -x[1]):
        print(f"  {status or '(blank)'}: {count}")
    print()

    # ── 2. Mapped rows breakdown by Source ──
    mapped_rows = [r for r in all_data if str(r.get("Manual User Check") or "").strip() == "Mapped"]
    print(f"2) MAPPED ROWS: {len(mapped_rows)}")
    print("-" * 50)

    source_counts = defaultdict(int)
    for row in mapped_rows:
        src = str(row.get("Source") or "").strip()
        source_counts[src] += 1

    for src, count in sorted(source_counts.items(), key=lambda x: -x[1]):
        print(f"  Source={src or '(blank)'}: {count}")
    print()

    # ── 3. Mapped rows breakdown by Source + 13WCF Line Item Mapping ──
    print("3) MAPPED ROWS BY SOURCE + LINE ITEM MAPPING")
    print("-" * 70)

    source_mapping_counts = defaultdict(int)
    for row in mapped_rows:
        src = str(row.get("Source") or "").strip()
        mapping = str(row.get("13WCF Line Item Mapping") or "").strip()
        source_mapping_counts[(src, mapping)] += 1

    for (src, mapping), count in sorted(source_mapping_counts.items(), key=lambda x: (-x[1], x[0])):
        print(f"  Source={src:<20s}  Mapping={mapping:<40s}  Count={count}")
    print()

    # ── 4. Intercompany Inflow / Outflow detail ──
    interco_rows = [
        r for r in all_data
        if str(r.get("13WCF Line Item Mapping") or "").strip() in ("Intercompany Inflow", "Intercompany Outflow")
    ]

    print(f"4) INTERCOMPANY INFLOW/OUTFLOW ROWS (all, not just mapped): {len(interco_rows)}")
    print("-" * 70)

    interco_by_source = defaultdict(int)
    interco_by_mapping = defaultdict(int)
    for row in interco_rows:
        src = str(row.get("Source") or "").strip()
        mapping = str(row.get("13WCF Line Item Mapping") or "").strip()
        interco_by_source[src] += 1
        interco_by_mapping[mapping] += 1

    print("  By Source:")
    for src, count in sorted(interco_by_source.items(), key=lambda x: -x[1]):
        print(f"    {src}: {count}")
    print("  By Mapping:")
    for mapping, count in sorted(interco_by_mapping.items(), key=lambda x: -x[1]):
        print(f"    {mapping}: {count}")
    print()

    # ── 5. Intercompany rows grouped by 13WCF Ref # ──
    print("5) INTERCOMPANY ROWS BY 13WCF Ref #")
    print("-" * 90)

    interco_ref_groups = defaultdict(list)
    for row in interco_rows:
        ref = str(row.get("13WCF Ref #") or "").strip()
        interco_ref_groups[ref].append(row)

    for ref, rows in sorted(interco_ref_groups.items()):
        total_usd = sum(float(r.get("Net Activity - USD") or 0) for r in rows)
        print(f"  Ref={ref}  Rows={len(rows)}  Sum(Net Activity - USD)={total_usd:,.2f}")
        for r in rows:
            acct = str(r.get("Account Name") or "")[:40]
            asset = str(r.get("Asset Code") or "")
            mapping = str(r.get("13WCF Line Item Mapping") or "")
            usd = r.get("Net Activity - USD") or 0
            rd = str(r.get("R/D") or "")
            check = str(r.get("Manual User Check") or "")
            counter = str(r.get("Counter Asset Base Code") or "")
            counter_exists = r.get("Counter_Asset_Exists")
            print(f"    Acct={acct:<40s}  Asset={asset:<8s}  Counter={counter:<8s}  "
                  f"USD={float(usd):>14,.2f}  R/D={rd:<12s}  Mapping={mapping:<25s}  Check={check}")
    print()

    # ── 6. All Lukka rows grouped by 13WCF Ref # — full picture for intercompany analysis ──
    lukka_rows = [r for r in all_data if str(r.get("Source") or "").strip() == "Lukka"]
    lukka_include = [r for r in lukka_rows if str(r.get("Incl/Excl") or "").strip() == "Include"]

    print(f"6) ALL LUKKA INCLUDE ROWS: {len(lukka_include)}")
    print("-" * 90)

    lukka_ref_groups = defaultdict(list)
    for row in lukka_include:
        ref = str(row.get("13WCF Ref #") or "").strip()
        if ref and ref != "0" and ref != "None":
            lukka_ref_groups[ref].append(row)

    # Find groups where unmapped rows sum ~0 but all rows don't (the bug scenario)
    print("  Ref groups where UNMAPPED sum differs significantly from ALL-rows sum:")
    print("  (These are groups affected by the 'only sum unmapped' change)")
    print()

    suspicious_groups = []
    for ref, rows in sorted(lukka_ref_groups.items()):
        sum_all = sum(float(r.get("Net Activity - USD") or 0) for r in rows)
        unmapped = [r for r in rows if not str(r.get("13WCF Line Item Mapping") or "").strip()]
        sum_unmapped = sum(float(r.get("Net Activity - USD") or 0) for r in unmapped)

        has_interco = any(
            str(r.get("13WCF Line Item Mapping") or "").strip() in ("Intercompany Inflow", "Intercompany Outflow")
            for r in rows
        )

        # Flag groups where the sums diverge meaningfully
        if abs(sum_all - sum_unmapped) > 0.01:
            suspicious_groups.append((ref, rows, sum_all, sum_unmapped, has_interco))

    if suspicious_groups:
        for ref, rows, sum_all, sum_unmapped, has_interco in suspicious_groups:
            marker = " *** HAS INTERCOMPANY ***" if has_interco else ""
            unmapped_near_zero = " [UNMAPPED ~0]" if abs(sum_unmapped) < 1 else ""
            all_near_zero = " [ALL ~0]" if abs(sum_all) < 1 else ""
            print(f"  Ref={ref}  #Rows={len(rows)}  "
                  f"Sum(all)={sum_all:>14,.2f}{all_near_zero}  "
                  f"Sum(unmapped)={sum_unmapped:>14,.2f}{unmapped_near_zero}"
                  f"{marker}")
            for r in rows:
                acct = str(r.get("Account Name") or "")[:35]
                asset = str(r.get("Asset Code") or "")
                mapping = str(r.get("13WCF Line Item Mapping") or "")
                usd = float(r.get("Net Activity - USD") or 0)
                check = str(r.get("Manual User Check") or "")
                matched = str(r.get("Matched_Substring") or "")[:30]
                print(f"      Acct={acct:<35s}  Asset={asset:<8s}  USD={usd:>14,.2f}  "
                      f"Mapping={mapping:<30s}  Check={check:<20s}  Match={matched}")
            print()
    else:
        print("  (none found)")
    print()

    # ── 7. Summary for quick comparison ──
    print("7) QUICK COMPARISON SUMMARY")
    print("=" * 50)
    print(f"  Target mapped rows:    {len(mapped_rows)}")
    print(f"  Our pipeline mapped:   431  (hardcoded for reference)")
    print(f"  Difference:            {431 - len(mapped_rows)}")
    print()
    print(f"  Target intercompany rows (Inflow+Outflow): {len(interco_rows)}")
    print()

    # Count mapped Lukka intercompany specifically
    lukka_interco_mapped = [
        r for r in mapped_rows
        if str(r.get("Source") or "").strip() == "Lukka"
        and str(r.get("13WCF Line Item Mapping") or "").strip() in ("Intercompany Inflow", "Intercompany Outflow")
    ]
    print(f"  Target mapped Lukka intercompany: {len(lukka_interco_mapped)}")
    print()

    # Not-mapped check rows
    not_mapped = [r for r in all_data if str(r.get("Manual User Check") or "").strip() == "Not Mapped - Check"]
    print(f"  Target 'Not Mapped - Check' rows: {len(not_mapped)}")
    print()

    # If difference is 2, the likely cause is 2 rows that our pipeline maps as
    # intercompany but the target leaves as Not Mapped - Check
    if 431 - len(mapped_rows) == 2:
        print("  ** The +2 difference likely means our pipeline is mapping 2 extra rows")
        print("     as Intercompany Inflow/Outflow that the target leaves unmapped.")
        print("     Check section 6 above for ref groups where unmapped-sum ~0 but all-sum != 0.")


if __name__ == "__main__":
    main()
