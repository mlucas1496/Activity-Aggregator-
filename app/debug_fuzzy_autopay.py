"""Debug: What fuzzy-matches the AUTOPAY OUT rows?"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

from pipeline.helpers import clean_notes_underscore
from pipeline.stage4_fuzzy import fuzzy_score, find_best_fuzzy_match
import openpyxl
import json

# Load supplementary historicals
ref_path = os.path.join(os.path.dirname(__file__), "pipeline", "wd_historical_mappings.json")
with open(ref_path) as f:
    supp = json.load(f)

# Load previous week historicals
PREV = "/Users/matthew.lucas/Downloads/Activity Aggregator - 02.27.2026 Forecast (2).xlsx"
print("Loading prev week historicals...")
wb = openpyxl.load_workbook(PREV, read_only=True, data_only=True)
ws = wb["Alteryx_Output"]
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_idx = {h: i for i, h in enumerate(headers) if h}

prev_historicals = []
for row in ws.iter_rows(min_row=2, values_only=True):
    source = row[col_idx.get("Source", -1)] or ""
    if source != "Workday":
        continue
    mapping = row[col_idx.get("13WCF Line Item Mapping", -1)] or ""
    if not mapping:
        continue
    acct = row[col_idx.get("Account Name", -1)] or ""
    notes = row[col_idx.get("Notes", -1)] or ""
    prev_historicals.append({
        "Account Name": acct,
        "Notes": str(notes),
        "Source": "Workday",
        "13WCF Line Item Mapping": mapping
    })
wb.close()

# Combine
all_historicals = prev_historicals + supp

# Filter for HSBC HKD Current Account
hsbc_hists = [h for h in all_historicals if "HSBC" in str(h.get("Account Name", "")) and "HKD Current" in str(h.get("Account Name", ""))]
print(f"\nHistoricals for HSBC HKD Current Account: {len(hsbc_hists)}")
for h in hsbc_hists:
    cn = clean_notes_underscore(h["Notes"])
    print(f"  Notes: {h['Notes'][:60]}")
    print(f"  Clean: {cn[:60]}")
    print(f"  Mapping: {h['13WCF Line Item Mapping']}")
    print()

# Test fuzzy matching against AUTOPAY
test_notes = [
    "/BTOT/1/SUPP/AUTOPAY OUT F02",
    "/BTOT/1/SUPP/AUTOPAY OUT F01",
]

# Build candidates
candidates = []
seen = set()
for h in hsbc_hists:
    cn = clean_notes_underscore(h["Notes"])
    if cn not in seen:
        candidates.append({"cleanNotes": cn, "mapping": h["13WCF Line Item Mapping"]})
        seen.add(cn)

print(f"Unique candidates: {len(candidates)}")
print()

for notes in test_notes:
    cn = clean_notes_underscore(notes)
    print(f"Testing: {notes}")
    print(f"  Clean: {cn}")
    print(f"  Tokens: {[t for t in cn.split('_') if t]}")

    # Score against each candidate
    for cand in candidates:
        score = fuzzy_score(cn, cand["cleanNotes"])
        if score > 0.3:
            print(f"  vs {cand['cleanNotes'][:60]} -> score={score:.3f} ({cand['mapping']})")

    result = find_best_fuzzy_match(cn, candidates)
    if result:
        print(f"  MATCH: {result['match']['mapping']} score={result['score']:.3f}")
        print(f"  Matched to: {result['match']['cleanNotes'][:60]}")
    else:
        print(f"  No match above threshold")
    print()
