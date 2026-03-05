"""
Stage 1: Load & Parse All Inputs
Reads input files + previous week output, builds lookup dicts from Excel sheets.

Sheet layouts (from actual files):
  - Most mapping sheets have blank leading rows/columns
  - Headers are at non-standard row offsets
  - Data columns are often offset by 1 (column A is blank)
"""
import csv
import io
import json
import os
from datetime import datetime as dt

import openpyxl
import pandas as pd

from config import BTH_NEEDED_COLS
from pipeline.helpers import date_to_serial, parse_date_str, serial_to_ymd


def load_inputs(file_paths, log):
    """
    Parse all uploaded files and build lookup maps.

    file_paths dict keys:
        prev_week, bank_statements, all_transactions, loan_report,
        search_strings, static_mapping
    """
    log("Stage 1: Parsing input files...")

    # ── 1. Previous Week's Activity Aggregator ──
    log("  Parsing previous week's Activity Aggregator...")
    prev_wb = openpyxl.load_workbook(file_paths["prev_week"], read_only=True, data_only=True)

    historicals = []
    known_lukka_accounts = set()

    if "Alteryx_Output" in prev_wb.sheetnames:
        ws = prev_wb["Alteryx_Output"]
        first_row = True
        for row in ws.iter_rows(values_only=True):
            if first_row:
                first_row = False
                continue
            if len(row) <= 23:
                continue
            source = _s(row[20])
            if not source:
                continue
            if source == "Lukka":
                acct = _s(row[0])
                if acct:
                    known_lukka_accounts.add(acct)
            mapping = _s(row[23])
            if not mapping or source != "Workday":
                continue
            historicals.append({
                "Account Name": _s(row[0]),
                "Notes": _s(row[3]),
                "Source": source,
                "13WCF Line Item Mapping": mapping,
            })

    # Supplement with reference WD historicals
    _load_supplementary_historicals(historicals, log)

    log(f"  -> Alteryx_Output: {len(historicals)} Workday historical rows with valid mappings")
    log(f"  -> Known Lukka accounts from prev week: {len(known_lukka_accounts)}")

    prev_sheet_names = [n for n in prev_wb.sheetnames if n != "Alteryx_Output"]

    # ── 2. Build lookup maps from prev week + Static_Mapping ──
    log("  Building lookup maps...")

    # Also open Static_Mapping for Bal Lukka Ref Mapping (primary source)
    sm_wb = openpyxl.load_workbook(file_paths["static_mapping"], read_only=True, data_only=True)

    # WD Bank acct -> Ledger acct Map (header row 4, data col 1)
    wd_bank_acct_map = _build_wd_bank_acct_map(prev_wb, sm_wb)
    log(f"  -> WD Bank Acct Map: {len(wd_bank_acct_map)} entries")

    # Wallet Mapping (header row 1, data col 0)
    wallet_map = _build_wallet_map(prev_wb, sm_wb)
    log(f"  -> Wallet Mapping: {len(wallet_map)} entries")

    # Lukka Ref Mapping — use "Bal Lukka Ref Mapping" from Static_Mapping (header row 0)
    lukka_ref_map = _build_lukka_ref_map(sm_wb)
    log(f"  -> Lukka Ref Mapping: {len(lukka_ref_map)} entries")

    # Legal Entity to Enterprise (header row 1, col 1)
    legal_entity_map = _build_legal_entity_map(prev_wb, sm_wb)
    log(f"  -> Legal Entity Map: {len(legal_entity_map)} entries")

    # Calendar (header row 7, col 1)
    calendar_map = _build_calendar_map(prev_wb, sm_wb)
    log(f"  -> Calendar Map: {len(calendar_map)} days")

    # Calendar Mapping (header row 3, col 1; weekEnding = EOWEEK col 10)
    calendar_mapping_map = _build_calendar_mapping_map(prev_wb, sm_wb)
    log(f"  -> Calendar Mapping: {len(calendar_mapping_map)} days")

    # FX (header row 9 in prev week, row 0 in static mapping)
    fx_map = _build_fx_map(prev_wb, sm_wb)
    log(f"  -> FX rates: {len(fx_map)} entries")

    # Mapping - Workday Accounts (header row 2, col 1; from prev week)
    wd_acct_flag_map = _build_wd_acct_flag_map(prev_wb)
    log(f"  -> WD Account Flag Map: {len(wd_acct_flag_map)} entries")

    sm_wb.close()

    # ── 3. Search Strings + Supplier/Customer ──
    log("  Loading Search_Strings_V2.xlsx...")
    ss_wb = openpyxl.load_workbook(file_paths["search_strings"], read_only=True, data_only=True)

    search_strings = _load_search_strings(ss_wb)
    log(f"  -> Search Strings: {len(search_strings)} mappings")

    # Supplier rows: from Search_Strings Entity_Supplier sheet or Static_Mapping
    supplier_rows = _load_supplier_rows(ss_wb, sm_path=file_paths["static_mapping"])
    log(f"  -> Entity_Supplier_13WCF_Mapping: {len(supplier_rows)} rows")

    # Customer rows: from Static_Mapping
    customer_rows = _load_customer_rows(sm_path=file_paths["static_mapping"])
    log(f"  -> Customer_13WCF_Mapping: {len(customer_rows)} rows")

    ss_wb.close()

    # ── 4. Bank Statements ──
    log("  Parsing bank statements...")
    bank_df = pd.read_excel(file_paths["bank_statements"], dtype=str)
    bank_rows = bank_df.to_dict("records")
    log(f"  -> Bank statement lines: {len(bank_rows)}")

    # ── 5. All Transactions CSV ──
    log("  Parsing All Transactions CSV (may take a moment)...")
    all_txns = _parse_csv_all_transactions(file_paths["all_transactions"])
    log(f"  -> All Transactions: {len(all_txns)} rows")

    # ── 6. LoanMasterReport CSV ──
    log("  Parsing LoanMasterReport CSV...")
    loan_rows = _parse_csv_loan_report(file_paths["loan_report"])
    log(f"  -> Loan rows: {len(loan_rows)}")

    log("Stage 1 complete -- all inputs parsed")

    return {
        "historicals": historicals,
        "known_lukka_accounts": known_lukka_accounts,
        "bank_rows": bank_rows,
        "all_txns": all_txns,
        "loan_rows": loan_rows,
        "search_strings": search_strings,
        "wd_bank_acct_map": wd_bank_acct_map,
        "wallet_map": wallet_map,
        "lukka_ref_map": lukka_ref_map,
        "legal_entity_map": legal_entity_map,
        "calendar_map": calendar_map,
        "calendar_mapping_map": calendar_mapping_map,
        "fx_map": fx_map,
        "supplier_rows": supplier_rows,
        "customer_rows": customer_rows,
        "wd_acct_flag_map": wd_acct_flag_map,
        "prev_wb": prev_wb,
        "prev_sheet_names": prev_sheet_names,
    }


# ── Utility helpers ──

def _s(v):
    """Safely convert to stripped string, return '' for None."""
    if v is None:
        return ""
    return str(v).strip()


def _num(v):
    """Convert to number, return value as-is if already numeric."""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return v
    try:
        f = float(str(v))
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return v


def _to_serial(date_val):
    """Convert datetime/int/string to Excel serial number."""
    if isinstance(date_val, (int, float)):
        return int(date_val)
    if isinstance(date_val, dt):
        return date_to_serial(date_val)
    s = str(date_val).strip()
    return parse_date_str(s)


def _find_header_row(ws, marker_col_value, max_rows=15):
    """Find the row index containing marker_col_value in any cell."""
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=max_rows)):
        for cell in row:
            if cell is not None and _s(cell) == marker_col_value:
                return i
    return None


def _read_sheet_at(wb, sheet_name, header_row, col_offset=0, max_rows=None):
    """
    Read a sheet as list-of-dicts, with header at header_row and data columns
    offset by col_offset from the left.
    Duplicate column names get a _2, _3 suffix to avoid overwrites.
    """
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True, max_row=max_rows))
    if header_row >= len(all_rows):
        return []
    header_raw = all_rows[header_row]
    headers = []
    seen = {}
    for i, h in enumerate(header_raw):
        if i < col_offset:
            continue
        name = _s(h) if h is not None else f"_col{i}"
        # Deduplicate column names
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        headers.append(name)

    result = []
    for row in all_rows[header_row + 1:]:
        obj = {}
        has_data = False
        for j, h in enumerate(headers):
            cell_idx = j + col_offset
            v = row[cell_idx] if cell_idx < len(row) else None
            if v is not None:
                has_data = True
            obj[h] = v
        if has_data:
            result.append(obj)
    return result


# ── Build lookup maps ──

def _build_wd_bank_acct_map(prev_wb, sm_wb):
    """
    WD Bank acct -> Ledger acct Map
    Layout: header at row 4, col 1: Bank Account, Entity, Ledger Account, CS Mapping - Entity, 13WCF Ref #
    Merge both sources — Static_Mapping has updated refs (e.g. ref=92 vs prev week's "RCF").
    Static_Mapping values take precedence.
    """
    m = {}
    for wb in [prev_wb, sm_wb]:
        rows = _read_sheet_at(wb, "WD Bank acct -> Ledger acct Map", header_row=4, col_offset=1)
        for r in rows:
            bank_acct = _s(r.get("Bank Account"))
            if not bank_acct:
                continue
            ref = r.get("13WCF Ref #")
            ref_str = str(int(ref)) if isinstance(ref, (int, float)) else (_s(ref) if ref is not None else "")
            m[bank_acct] = {
                "entity": _s(r.get("Entity")),
                "ref": ref_str,
                "csEntity": _s(r.get("CS Mapping - Entity")),
            }
    return m


def _build_wallet_map(prev_wb, sm_wb):
    """
    Wallet Mapping
    Layout: header at row 1, col 0: Wallet Name, Wallet ID, Enterprise, HOT/WARM/COLD, Custodial/House, Vendor
    Merge both sources. Process prev_wb first, then sm_wb (Static_Mapping takes precedence).
    Within a sheet, later rows overwrite earlier ones (handles duplicate wallet names with
    updated enterprise/vendor values needed for Lukka Ref Mapping key composition).
    """
    m = {}
    for wb in [prev_wb, sm_wb]:
        rows = _read_sheet_at(wb, "Wallet Mapping", header_row=1, col_offset=0)
        for r in rows:
            wallet = _s(r.get("Wallet Name"))
            if not wallet:
                continue
            m[wallet] = {
                "enterprise": _s(r.get("Enterprise")),
                "custodialHouse": _s(r.get("Custodial/House") or r.get("House/Custodial")),
                "vendor": _s(r.get("Vendor")),
            }
    return m


def _build_lukka_ref_map(sm_wb):
    """
    Bal Lukka Ref Mapping (from Static_Mapping)
    Layout: header at row 0, col 0:
      Account Name, Sub-Account Name, Asset Code, Enterprise, House/Custodial, Vendor, Account Type, Ref ID, Ref
    Key = Ref ID column (pre-built composite key), Value = Ref number
    """
    rows = _read_sheet_at(sm_wb, "Bal Lukka Ref Mapping", header_row=0, col_offset=0)
    m = {}
    for r in rows:
        ref_id = _s(r.get("Ref ID"))
        ref = r.get("Ref")
        if ref_id and ref is not None:
            m[ref_id] = str(int(ref)) if isinstance(ref, (int, float)) else _s(ref)
    return m


def _build_legal_entity_map(prev_wb, sm_wb):
    """
    Legal Entity to Enterprise
    Layout: header at row 1, col 1: LOOKUP, Legal Entity, (blank), Enterprise, Account Type
    JS maps: LOOKUP -> {enterprise, accountType} AND Legal Entity -> same
    """
    for wb in [prev_wb, sm_wb]:
        rows = _read_sheet_at(wb, "Legal Entity to Enterprise", header_row=1, col_offset=1)
        if rows:
            break
    m = {}
    for r in rows:
        lookup = _s(r.get("LOOKUP"))
        legal_entity = _s(r.get("Legal Entity"))
        enterprise = _s(r.get("Enterprise"))
        account_type = _s(r.get("Account Type"))
        obj = {"legalEntity": legal_entity, "enterprise": enterprise, "accountType": account_type}
        if lookup:
            m[lookup] = obj
        if legal_entity:
            m[legal_entity] = obj
    return m


def _build_calendar_map(prev_wb, sm_wb):
    """
    Calendar
    Layout: header at row 7, col 1: Day, Activity Week, Actuals Week Beginning, Actuals Week End
    Key = date serial, Value = {activityWeek, actualsWeekBeg, actualsWeekEnd}
    Merge both sources — Static_Mapping has wider range (to 2027), prev week may be shorter.
    """
    m = {}
    for wb in [sm_wb, prev_wb]:
        rows = _read_sheet_at(wb, "Calendar", header_row=7, col_offset=1)
        for r in rows:
            day = r.get("Day")
            if day is None:
                continue
            serial = _to_serial(day)
            if serial is None:
                continue
            if serial not in m:
                m[serial] = {
                    "activityWeek": _num(r.get("Activity Week")),
                    "actualsWeekBeg": _num(r.get("Actuals Week Beginning")),
                    "actualsWeekEnd": _num(r.get("Actuals Week End")),
                }
    return m


def _build_calendar_mapping_map(prev_wb, sm_wb):
    """
    Calendar Mapping
    Layout: header at row 3, col 1: Date, Day, Week Within Month, Week, Month, Weekday,
            Weeks in Month, Year, BEGWEEK, EOWEEK, EOMONTH, ...
    weekEnding = EOWEEK (col 10, first occurrence — dedup makes second one EOWEEK_2)
    Merge both sources for wider date coverage.
    """
    m = {}
    for wb in [sm_wb, prev_wb]:
        rows = _read_sheet_at(wb, "Calendar Mapping", header_row=3, col_offset=1)
        for r in rows:
            date_val = r.get("Date")
            if date_val is None:
                continue
            serial = _to_serial(date_val)
            if serial is None:
                continue
            if serial in m:
                continue
            # EOWEEK is the week ending date — convert to serial
            eoweek = r.get("EOWEEK")
            week_ending = ""
            if eoweek is not None:
                we_serial = _to_serial(eoweek)
                if we_serial is not None:
                    week_ending = we_serial
            m[serial] = {
                "weekEnding": week_ending,
                "week": _num(r.get("Week")),
                "month": _num(r.get("Month")),
                "year": _num(r.get("Year")),
            }
    return m


def _build_fx_map(prev_wb, sm_wb):
    """
    FX rates
    Prev week layout: header at row 9, col 1: Date, Open, High, Low, Close, Currency
    Static mapping layout: header at row 0, col 0: Date, Open, High, Low, Close, Currency
    Key = "YYYY-MM-DD|CCY", Value = Close price
    """
    m = {}
    # Try static mapping first (simpler layout)
    rows = _read_sheet_at(sm_wb, "FX", header_row=0, col_offset=0)
    if rows:
        _parse_fx_rows(rows, m)
    # Also try prev week (may have more recent data)
    rows2 = _read_sheet_at(prev_wb, "FX", header_row=9, col_offset=1)
    if rows2:
        _parse_fx_rows(rows2, m)
    return m


def _parse_fx_rows(rows, m):
    """Parse FX rows into map. Handles column name variations (leading spaces, etc)."""
    for r in rows:
        date_val = r.get("Date") or r.get("date")
        # Handle column names with leading spaces from prev week
        close = r.get("Close") or r.get(" Close") or r.get("close")
        ccy = r.get("Currency") or r.get("currency")
        if date_val is None or close is None or ccy is None:
            continue
        serial = _to_serial(date_val)
        if serial is None:
            continue
        ymd = serial_to_ymd(serial)
        if not ymd:
            continue
        ccy_str = _s(ccy)
        try:
            rate = float(close)
        except (ValueError, TypeError):
            continue
        key = f"{ymd}|{ccy_str}"
        if key not in m:
            m[key] = rate


def _build_wd_acct_flag_map(prev_wb):
    """
    Mapping - Workday Accounts (from prev week)
    Layout: header at row 2, col 1: Bank Account, Consol. Entity, Include/Exclude, 13WCF Ref #
    """
    rows = _read_sheet_at(prev_wb, "Mapping - Workday Accounts", header_row=2, col_offset=1)
    m = {}
    for r in rows:
        bank_acct = _s(r.get("Bank Account"))
        if not bank_acct:
            continue
        ref = r.get("13WCF Ref #")
        ref_str = str(int(ref)) if isinstance(ref, (int, float)) else _s(ref)
        m[bank_acct] = {
            "consolEntity": _s(r.get("Consol. Entity")),
            "inclExcl": _s(r.get("Include/Exclude")),
            "ref": ref_str,
        }
    return m


# ── Search Strings + Supplier/Customer ──

def _load_search_strings(ss_wb):
    """
    Search_Strings_V2 -> 'Mappings - New' sheet
    Layout: header at row 0: Substrings, R/D, Consolidated Entity, 13WCF Line Item Mapping
    """
    for name in ["Mappings - New", "Mappings", "Sheet1"]:
        if name in ss_wb.sheetnames:
            return _read_sheet_at(ss_wb, name, header_row=0, col_offset=0)
    if ss_wb.sheetnames:
        return _read_sheet_at(ss_wb, ss_wb.sheetnames[0], header_row=0, col_offset=0)
    return []


def _load_supplier_rows(ss_wb, sm_path):
    """
    Load suppliers. The JS uses Static_Mapping format:
      Bullish Group Hierarchy, Supplier, 13WCF Mapping, Entity, R/D
    But Search_Strings also has an Entity_Supplier sheet with different columns:
      Bullish Ledger Accounts, Bullish Group Hierarchy, 13WCF Mapped Entity, Supplier, Spend Category, 13WCF Mapping
    We read from Static_Mapping (matches the JS embedded format).
    """
    sm_wb2 = openpyxl.load_workbook(sm_path, read_only=True, data_only=True)
    rows = _read_sheet_at(sm_wb2, "Entity_Supplier_13WCF_Mapping", header_row=0, col_offset=0)
    sm_wb2.close()
    # Normalize to JS format: Supplier, 13WCF Mapping, R/D
    result = []
    for r in rows:
        result.append({
            "Bullish Group Hierarchy": _s(r.get("Bullish Group Hierarchy")),
            "Supplier": _s(r.get("Supplier")),
            "13WCF Mapping": _s(r.get("13WCF Mapping")),
            "Entity": _s(r.get("Entity")),
            "R/D": _s(r.get("R/D")),
        })
    return result


def _load_customer_rows(sm_path):
    """
    Customer_13WCF_Mapping from Static_Mapping
    Layout: header at row 0: Sold-To Customer, 13WCF Mapping, 13WCF Entity, R/D
    """
    sm_wb2 = openpyxl.load_workbook(sm_path, read_only=True, data_only=True)
    rows = _read_sheet_at(sm_wb2, "Customer_13WCF_Mapping", header_row=0, col_offset=0)
    sm_wb2.close()
    return rows


# ── CSV parsing ──

def _parse_csv_all_transactions(path):
    """Parse the All Transactions CSV (skip first 5 metadata rows, row 6 is header)."""
    rows = []
    with open(path, "r", encoding="utf-8-sig") as f:
        for _ in range(5):
            next(f, None)
        reader = csv.DictReader(f)
        for raw in reader:
            slim = {}
            for key in BTH_NEEDED_COLS:
                v = raw.get(key)
                if v is not None and v != "":
                    slim[key] = v
            rows.append(slim)
    return rows


def _parse_csv_loan_report(path):
    """Parse the LoanMasterReport CSV."""
    rows = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(dict(row))
    return rows


def _load_supplementary_historicals(historicals, log):
    """Load reference WD historicals from wd_historical_mappings.json."""
    ref_path = os.path.join(os.path.dirname(__file__), "wd_historical_mappings.json")
    if not os.path.exists(ref_path):
        return
    try:
        with open(ref_path, "r") as f:
            entries = json.load(f)
        # Deduplicate against existing historicals
        existing = {(h["Account Name"], h["Notes"]) for h in historicals}
        added = 0
        for entry in entries:
            key = (entry["Account Name"], entry["Notes"])
            if key not in existing:
                historicals.append(entry)
                existing.add(key)
                added += 1
        if added:
            log(f"  -> Supplementary WD historicals: {added} added from reference file")
    except Exception as e:
        log(f"  Warning: Could not load supplementary historicals: {e}")
