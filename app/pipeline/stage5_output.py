"""
Stage 5: Manual User Check + Write Output XLSX
Writes multi-sheet XLSX using xlsxwriter (constant_memory mode for large sheets).
Formatting matches the Alteryx target output.
"""
import os
import math
from datetime import datetime
import xlsxwriter
import openpyxl

from config import ALTERYX_COLS, BTH_COLS, BTH_ROW8_SPECIAL, BTH_KEY_MAP, WD_COLS, COPY_SHEET_NAMES


# ── Theme color approximations (Office default theme) ──
_THEME2_BG = "#E7E6E6"      # light gray (Workday header fill)
_THEME3_T9_BG = "#ECEEF0"   # light blue-gray (A&M header fill, theme=3 tint 0.9)


def generate_output(all_rows, bth_rows, wd_rows, prev_wb, prev_sheet_names, prev_week_path, output_dir, log):
    """
    Apply Manual User Check and write multi-sheet XLSX.
    Returns dict with {file_path, file_name, stats}.
    """
    log("Stage 5: Generating output...")

    # Apply Manual User Check
    mapped_count = 0
    not_mapped_count = 0
    exclude_count = 0

    for row in all_rows:
        mapping = row.get("13WCF Line Item Mapping") or ""
        incl_excl = row.get("Inc/Excl") or row.get("Incl/Excl") or ""

        if not mapping and incl_excl == "Include":
            row["Manual User Check"] = "Not Mapped - Check"
            not_mapped_count += 1
        elif incl_excl == "Exclude":
            row["Manual User Check"] = "Exclude - Pass"
            exclude_count += 1
        else:
            row["Manual User Check"] = "Mapped"
            mapped_count += 1

        # Normalize for Alteryx_Output column "Incl/Excl"
        row["Incl/Excl"] = row.get("Inc/Excl") or incl_excl
        if not row.get("Entity/Enterprise"):
            row["Entity/Enterprise"] = row.get("Entity") or row.get("Enterprise") or ""

    log(f"  -> Mapped: {mapped_count}, Not Mapped: {not_mapped_count}, Excluded: {exclude_count}")

    # Build output file
    now = datetime.now()
    file_name = f"Activity Aggregator - {now.strftime('%m.%d.%Y')} Forecast.xlsx"
    file_path = os.path.join(output_dir, file_name)

    # Use xlsxwriter with constant_memory for large sheets
    workbook = xlsxwriter.Workbook(file_path, {"constant_memory": True, "nan_inf_to_errors": True})

    # ── Build reusable formats ──
    fmts = _build_formats(workbook)

    # Re-open previous week file in full mode for formatted copying
    log("  Re-opening previous week file for formatted sheet copying...")
    prev_wb_full = openpyxl.load_workbook(prev_week_path, data_only=True)

    # ── Copy pre-data separator sheets ──
    for sheet_name in ["Update Instructions", "Alteryx Column Agg Mapping", "Alteryx Output ->"]:
        if sheet_name in prev_sheet_names:
            try:
                _copy_sheet(workbook, prev_wb_full, sheet_name)
            except Exception as e:
                log(f"  Warning: Could not copy sheet '{sheet_name}': {e}")

    # ── Alteryx_Output ──
    log("  Writing Alteryx_Output sheet...")
    _write_alteryx_output(workbook, fmts, all_rows, log)

    # ── Copy "Alteryx Inputs ->" separator ──
    if "Alteryx Inputs ->" in prev_sheet_names:
        try:
            _copy_sheet(workbook, prev_wb_full, "Alteryx Inputs ->")
        except Exception:
            pass

    # ── WD - Find Bank Statement Lines ──
    log("  Writing WD - Find Bank Statement Lines sheet...")
    _write_wd_sheet(workbook, fmts, wd_rows, log)

    # ── BTH Trade Log ──
    log("  Writing BTH Trade Log sheet...")
    _write_bth_sheet(workbook, fmts, bth_rows, log)

    # ── Copy post-data mapping sheets ──
    log("  Copying mapping sheets forward...")
    post_data_sheets = [
        "Lukka Mapping ->",
        "Wallet Mapping", "Lukka Ref Mapping", "Legal Entity to Enterprise",
        "WD Mapping ->",
        "WD Bank acct -> Ledger acct Map", "Calendar", "Calendar Mapping",
        "FX", "Mapping - Workday Accounts",
    ]
    for sheet_name in post_data_sheets:
        if sheet_name in prev_sheet_names:
            try:
                _copy_sheet(workbook, prev_wb_full, sheet_name)
            except Exception as e:
                log(f"  Warning: Could not copy sheet '{sheet_name}': {e}")

    workbook.close()

    # Close workbooks
    try:
        prev_wb_full.close()
    except Exception:
        pass
    try:
        prev_wb.close()
    except Exception:
        pass

    stats = {
        "total": len(all_rows),
        "mapped": mapped_count,
        "notMapped": not_mapped_count,
        "excluded": exclude_count,
        "bthRows": len(bth_rows),
        "wdRows": len(wd_rows),
    }

    log(f"Stage 5 complete -- output: {file_name}")
    log(f"  Total rows: {stats['total']}")
    log(f"  Mapped: {stats['mapped']} ({(stats['mapped'] / max(stats['total'], 1) * 100):.1f}%)")
    log(f"  Not Mapped: {stats['notMapped']}")
    log(f"  Excluded: {stats['excluded']}")

    return {"file_path": file_path, "file_name": file_name, "stats": stats}


# ═══════════════════════════════════════════════════════════════════════
# Format factory
# ═══════════════════════════════════════════════════════════════════════

def _build_formats(wb):
    """Create all reusable xlsxwriter formats."""
    base = {"font_name": "Arial", "font_size": 10}

    f = {}

    # Alteryx_Output
    f["alt_default"] = wb.add_format({**base})
    f["alt_date_mdy_hm"] = wb.add_format({**base, "num_format": "m/d/yy h:mm"})
    f["alt_date_mmddyy"] = wb.add_format({**base, "num_format": "mm-dd-yy"})
    f["alt_fuzzy"] = wb.add_format({**base, "bg_color": "#FFFF00"})
    f["alt_fuzzy_date_mdy_hm"] = wb.add_format({**base, "bg_color": "#FFFF00", "num_format": "m/d/yy h:mm"})
    f["alt_fuzzy_date_mmddyy"] = wb.add_format({**base, "bg_color": "#FFFF00", "num_format": "mm-dd-yy"})

    # WD sheet
    f["wd_default"] = wb.add_format({**base})
    f["wd_source_label"] = wb.add_format({
        **base, "italic": True, "bg_color": "#FFFFFF", "align": "center",
    })
    f["wd_hdr_wd"] = wb.add_format({
        **base, "bold": True, "font_color": "#000000",
        "bg_color": _THEME2_BG, "align": "center", "valign": "top", "text_wrap": True,
    })
    f["wd_hdr_wd_left"] = wb.add_format({
        **base, "bold": True, "font_color": "#000000",
        "bg_color": _THEME2_BG, "align": "left", "valign": "top", "text_wrap": True,
    })
    f["wd_hdr_am"] = wb.add_format({
        **base, "bold": True, "font_color": "#000000",
        "bg_color": _THEME3_T9_BG, "align": "center", "valign": "vcenter", "text_wrap": True,
    })
    f["wd_date_mmddyy"] = wb.add_format({**base, "num_format": "mm-dd-yy"})
    f["wd_datetime"] = wb.add_format({**base, "num_format": "m/d/yy h:mm"})

    # BTH sheet
    f["bth_default"] = wb.add_format({**base})
    f["bth_date_mmddyy"] = wb.add_format({**base, "num_format": "mm-dd-yy"})
    f["bth_subtotal"] = wb.add_format({**base, "num_format": "#,##0.00"})

    return f


# ═══════════════════════════════════════════════════════════════════════
# Alteryx_Output sheet
# ═══════════════════════════════════════════════════════════════════════

def _write_alteryx_output(workbook, fmts, all_rows, log):
    ws = workbook.add_worksheet("Alteryx_Output")
    ws.set_default_row(12.75)

    # Column default formats
    for c in range(len(ALTERYX_COLS)):
        if c == 1:  # Transaction Date
            ws.set_column(c, c, None, fmts["alt_date_mdy_hm"])
        elif c == 6:  # Week Ending
            ws.set_column(c, c, None, fmts["alt_date_mmddyy"])
        else:
            ws.set_column(c, c, None, fmts["alt_default"])

    # Header row
    ws.set_row(0, 12.75)
    for c, col_name in enumerate(ALTERYX_COLS):
        ws.write_string(0, c, col_name, fmts["alt_default"])

    # Data rows
    for r_idx, row in enumerate(all_rows):
        excel_row = r_idx + 1
        is_fuzzy = str(row.get("Matched_Substring") or "").startswith("Fuzzy")

        for c, col_name in enumerate(ALTERYX_COLS):
            val = row.get(col_name)

            # Pick format: date column + fuzzy combination
            if c == 1:
                fmt = fmts["alt_fuzzy_date_mdy_hm"] if is_fuzzy else None  # None = column default
            elif c == 6:
                fmt = fmts["alt_fuzzy_date_mmddyy"] if is_fuzzy else None
            else:
                fmt = fmts["alt_fuzzy"] if is_fuzzy else None

            if val is not None and val != "":
                _write_cell(ws, excel_row, c, val, fmt)
            elif is_fuzzy:
                ws.write_blank(excel_row, c, None, fmts["alt_fuzzy"])


# ═══════════════════════════════════════════════════════════════════════
# WD - Find Bank Statement Lines sheet
# ═══════════════════════════════════════════════════════════════════════

def _write_wd_sheet(workbook, fmts, wd_rows, log):
    ws = workbook.add_worksheet("WD - Find Bank Statement Lines")

    # Sheet settings
    ws.set_zoom(70)

    # Column widths (xlsxwriter col index)
    ws.set_column(0, 0, 0.86)    # A - narrow
    ws.set_column(1, 1, 8.73)    # B
    ws.set_column(4, 4, 11.40)   # E - Statement Line Date
    ws.set_column(5, 5, 8.73)    # F
    ws.set_column(10, 10, 90.13) # K - Addenda (wide)
    ws.set_column(11, 11, 8.73)  # L
    ws.set_column(13, 13, 14.13) # N - Last Updated Date Time
    ws.set_column(14, 14, 8.73)  # O
    ws.set_column(16, 16, 13.40) # Q - Net Activity - USD
    ws.set_column(17, 17, 8.73)  # R
    ws.set_column(18, 18, 13.40) # S - Week Ending
    ws.set_column(19, 19, 8.73)  # T

    # Row 3: title (height 13.15)
    ws.set_row(3, 13.15)
    ws.write(3, 0, "Data Source(s): Workday Find Bank Statement Lines Report", fmts["wd_default"])

    # Row 4: period (height 13.15)
    ws.set_row(4, 13.15)
    ws.write(4, 0, "Period: ", fmts["wd_default"])

    # Row 6: source labels (italic, white fill, centered)
    for c in range(1, 14):
        ws.write(6, c, "Workday", fmts["wd_source_label"])
    for c in range(14, 27):
        ws.write(6, c, "A&M", fmts["wd_source_label"])

    # Row 7: main header row (height 52.5, formatted)
    ws.set_row(7, 52.5)
    # Workday section (cols 1-13 → WD_COLS[0..12])
    for c in range(13):
        col_idx = c + 1
        # Statement Line Date (col 4) is left-aligned
        hdr_fmt = fmts["wd_hdr_wd_left"] if col_idx == 4 else fmts["wd_hdr_wd"]
        ws.write(7, col_idx, WD_COLS[c], hdr_fmt)
    # A&M section (cols 14-26 → WD_COLS[13..25])
    for c in range(13, 26):
        col_idx = c + 1
        ws.write(7, col_idx, WD_COLS[c], fmts["wd_hdr_am"])

    # Row 8: sub-header (plain repeat)
    for c, col_name in enumerate(WD_COLS):
        ws.write(8, c + 1, col_name, fmts["wd_default"])

    # Auto-filter: row 7 through last data row, cols A-AA
    last_data_row = 8 + len(wd_rows)
    ws.autofilter(7, 0, last_data_row, 26)

    # Data rows starting at row 9
    for r_idx, row in enumerate(wd_rows):
        excel_row = r_idx + 9
        for c, col_name in enumerate(WD_COLS):
            col_idx = c + 1
            val = row.get(col_name)
            if val is None or val == "":
                continue
            # Date format for specific columns
            if c == 3:    # Statement Line Date → mm-dd-yy
                _write_cell(ws, excel_row, col_idx, val, fmts["wd_date_mmddyy"])
            elif c == 12:  # Last Updated Date Time → m/d/yy h:mm
                _write_cell(ws, excel_row, col_idx, val, fmts["wd_datetime"])
            elif c == 17:  # Week Ending → mm-dd-yy
                _write_cell(ws, excel_row, col_idx, val, fmts["wd_date_mmddyy"])
            else:
                _write_cell(ws, excel_row, col_idx, val)


# ═══════════════════════════════════════════════════════════════════════
# BTH Trade Log sheet
# ═══════════════════════════════════════════════════════════════════════

def _write_bth_sheet(workbook, fmts, bth_rows, log):
    ws = workbook.add_worksheet("BTH Trade Log")

    # Column widths
    ws.set_column(4, 4, 36.60)    # E - Sub Account Name
    ws.set_column(16, 16, 22.13)  # Q - Base Asset Amount
    ws.set_column(17, 17, 19.40)  # R - Price
    ws.set_column(20, 20, 12.73)  # U - Counter Asset Amount

    # Row 8 (0-indexed): partial header with subtotals
    for c in range(min(38, len(BTH_COLS))):
        ws.write(8, c, BTH_COLS[c], fmts["bth_default"])
    for col_idx, label in BTH_ROW8_SPECIAL.items():
        ws.write(8, col_idx, label, fmts["bth_default"])

    # Row 9: full header
    for c, col_name in enumerate(BTH_COLS):
        ws.write(9, c, col_name, fmts["bth_default"])

    # Auto-filter: row 9 through last data row
    last_data_row = 9 + len(bth_rows)
    ws.autofilter(9, 0, last_data_row, len(BTH_COLS) - 1)

    # Data rows starting at row 10
    for r_idx, row in enumerate(bth_rows):
        excel_row = r_idx + 10
        for c, col_name in enumerate(BTH_COLS):
            data_key = BTH_KEY_MAP.get(col_name, col_name)
            val = row.get(data_key)
            if val is None or val == "":
                continue
            # Date format for specific columns
            if c == 10:   # Transaction Date → mm-dd-yy
                _write_cell(ws, excel_row, c, val, fmts["bth_date_mmddyy"])
            elif c == 48:  # Week Ending → mm-dd-yy
                _write_cell(ws, excel_row, c, val, fmts["bth_date_mmddyy"])
            else:
                _write_cell(ws, excel_row, c, val)


# ═══════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════

def _write_cell(ws, row, col, val, fmt=None):
    """Write a value to a cell, choosing the right xlsxwriter method."""
    if isinstance(val, bool):
        ws.write_boolean(row, col, val, fmt)
    elif isinstance(val, (int, float)):
        if math.isnan(val) or math.isinf(val):
            ws.write_string(row, col, "", fmt)
        else:
            ws.write_number(row, col, val, fmt)
    elif isinstance(val, datetime):
        ws.write_datetime(row, col, val, fmt)
    else:
        ws.write_string(row, col, str(val), fmt)


# ═══════════════════════════════════════════════════════════════════════
# Sheet copying with formatting preservation
# ═══════════════════════════════════════════════════════════════════════

# Office 2016 default theme base colors (index → RRGGBB)
_THEME_COLORS = [
    "FFFFFF", "000000", "E7E6E6", "44546A",
    "4472C4", "ED7D31", "A5A5A5", "FFC000",
    "5B9BD5", "70AD47",
]

# Standard Excel indexed color palette (index → RRGGBB)
_INDEXED_COLORS = {
    0: "000000", 1: "FFFFFF", 2: "FF0000", 3: "00FF00",
    4: "0000FF", 5: "FFFF00", 6: "FF00FF", 7: "00FFFF",
    8: "000000", 9: "FFFFFF", 10: "FF0000", 11: "00FF00",
    12: "0000FF", 13: "FFFF00", 14: "FF00FF", 15: "00FFFF",
    16: "800000", 17: "008000", 18: "000080", 19: "808000",
    20: "800080", 21: "008080", 22: "C0C0C0", 23: "808080",
    24: "9999FF", 25: "993366", 26: "FFFFCC", 27: "CCFFFF",
    28: "660066", 29: "FF8080", 30: "0066CC", 31: "CCCCFF",
    32: "000080", 33: "FF00FF", 34: "FFFF00", 35: "00FFFF",
    36: "800080", 37: "800000", 38: "008080", 39: "0000FF",
    40: "00CCFF", 41: "CCFFFF", 42: "CCFFCC", 43: "FFFF99",
    44: "99CCFF", 45: "FF99CC", 46: "CC99FF", 47: "FFCC99",
    48: "3366FF", 49: "33CCCC", 50: "99CC00", 51: "FFCC00",
    52: "FF9900", 53: "FF6600", 54: "666699", 55: "969696",
    56: "003366", 57: "339966", 58: "003300", 59: "333300",
    60: "993300", 61: "993366", 62: "333399", 63: "333333",
    64: "000000",  # system foreground
}

# openpyxl border style → xlsxwriter border index
_BORDER_MAP = {
    "thin": 1, "medium": 2, "dashed": 3, "dotted": 4,
    "thick": 5, "double": 6, "hair": 7, "mediumDashed": 8,
    "dashDot": 9, "mediumDashDot": 10, "dashDotDot": 11,
    "mediumDashDotDot": 12, "slantDashDot": 13,
}


def _resolve_color(color):
    """Convert openpyxl Color → '#RRGGBB' or None."""
    if color is None:
        return None
    try:
        if color.type == "rgb":
            rgb = str(color.rgb)
            if rgb and rgb not in ("00000000", "0"):
                hex6 = rgb[2:] if len(rgb) == 8 else rgb
                if len(hex6) == 6:
                    return f"#{hex6}"
        elif color.type == "theme" and color.theme is not None:
            idx = color.theme
            if 0 <= idx < len(_THEME_COLORS):
                base = _THEME_COLORS[idx]
                tint = color.tint or 0
                if abs(tint) < 0.001:
                    return f"#{base}"
                r = int(base[0:2], 16)
                g = int(base[2:4], 16)
                b = int(base[4:6], 16)
                if tint > 0:
                    r = int(r + tint * (255 - r))
                    g = int(g + tint * (255 - g))
                    b = int(b + tint * (255 - b))
                else:
                    r = int(r * (1 + tint))
                    g = int(g * (1 + tint))
                    b = int(b * (1 + tint))
                return f"#{min(max(r,0),255):02X}{min(max(g,0),255):02X}{min(max(b,0),255):02X}"
        elif color.type == "indexed" and color.indexed is not None:
            hex6 = _INDEXED_COLORS.get(color.indexed)
            if hex6:
                return f"#{hex6}"
    except Exception:
        pass
    return None


def _cell_format_key(cell):
    """Extract a hashable tuple of formatting properties from an openpyxl cell."""
    font = cell.font
    fill = cell.fill
    border = cell.border
    align = cell.alignment
    numfmt = cell.number_format

    # Font
    f_name = font.name or ""
    f_size = font.sz or 11
    f_bold = bool(font.bold)
    f_italic = bool(font.italic)
    f_underline = font.underline or ""
    f_color = _resolve_color(font.color) if font.color else None
    f_strike = bool(font.strike)

    # Fill
    bg = None
    if fill.patternType == "solid" and fill.fgColor:
        bg = _resolve_color(fill.fgColor)

    # Borders
    bl = _BORDER_MAP.get(border.left.style, 0) if border.left and border.left.style else 0
    br = _BORDER_MAP.get(border.right.style, 0) if border.right and border.right.style else 0
    bt = _BORDER_MAP.get(border.top.style, 0) if border.top and border.top.style else 0
    bb = _BORDER_MAP.get(border.bottom.style, 0) if border.bottom and border.bottom.style else 0
    blc = _resolve_color(border.left.color) if border.left and border.left.color and bl else None
    brc = _resolve_color(border.right.color) if border.right and border.right.color and br else None
    btc = _resolve_color(border.top.color) if border.top and border.top.color and bt else None
    bbc = _resolve_color(border.bottom.color) if border.bottom and border.bottom.color and bb else None

    # Alignment
    ah = align.horizontal or ""
    av = align.vertical or ""
    aw = bool(align.wrap_text)
    ai = align.indent or 0

    return (
        f_name, f_size, f_bold, f_italic, f_underline, f_color, f_strike,
        bg,
        bl, br, bt, bb, blc, brc, btc, bbc,
        ah, av, aw, ai,
        numfmt or "General",
    )


def _make_format(workbook, key):
    """Create an xlsxwriter Format from a format-key tuple."""
    (f_name, f_size, f_bold, f_italic, f_underline, f_color, f_strike,
     bg,
     bl, br, bt, bb, blc, brc, btc, bbc,
     ah, av, aw, ai,
     numfmt) = key

    props = {}

    if f_name:
        props["font_name"] = f_name
    if f_size and f_size != 11:
        props["font_size"] = f_size
    if f_bold:
        props["bold"] = True
    if f_italic:
        props["italic"] = True
    if f_underline:
        props["underline"] = True
    if f_color:
        props["font_color"] = f_color
    if f_strike:
        props["font_strikeout"] = True

    if bg:
        props["bg_color"] = bg

    if bl:
        props["left"] = bl
    if br:
        props["right"] = br
    if bt:
        props["top"] = bt
    if bb:
        props["bottom"] = bb
    if blc:
        props["left_color"] = blc
    if brc:
        props["right_color"] = brc
    if btc:
        props["top_color"] = btc
    if bbc:
        props["bottom_color"] = bbc

    _VALIGN = {"top": "top", "center": "vcenter", "bottom": "bottom"}
    if ah:
        props["align"] = ah
    if av:
        props["valign"] = _VALIGN.get(av, av)
    if aw:
        props["text_wrap"] = True
    if ai:
        props["indent"] = ai

    if numfmt and numfmt != "General":
        props["num_format"] = numfmt

    if not props:
        return None
    return workbook.add_format(props)


# Default key for cells with no custom formatting
_DEFAULT_KEY = ("", 11, False, False, "", None, False,
                None,
                0, 0, 0, 0, None, None, None, None,
                "", "", False, 0,
                "General")


def _copy_sheet(workbook, prev_wb, sheet_name):
    """Copy a sheet from the previous openpyxl workbook, preserving formatting."""
    if sheet_name not in prev_wb.sheetnames:
        return
    src_ws = prev_wb[sheet_name]
    dst_ws = workbook.add_worksheet(sheet_name)

    # Format cache to avoid creating duplicate Format objects
    fmt_cache = {}

    # Copy column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width and dim.width > 0:
            from openpyxl.utils import column_index_from_string
            col_idx = column_index_from_string(col_letter) - 1
            dst_ws.set_column(col_idx, col_idx, dim.width)
        if dim.hidden:
            from openpyxl.utils import column_index_from_string
            col_idx = column_index_from_string(col_letter) - 1
            dst_ws.set_column(col_idx, col_idx, None, None, {"hidden": True})

    # Collect merged cell ranges
    merged_cells = set()
    for merged_range in src_ws.merged_cells.ranges:
        merged_cells.add((merged_range.min_row, merged_range.min_col,
                          merged_range.max_row, merged_range.max_col))

    # Track which cells are part of a merge (but not the top-left)
    merge_skip = set()
    for (r1, c1, r2, c2) in merged_cells:
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                if rr != r1 or cc != c1:
                    merge_skip.add((rr, cc))

    for row in src_ws.iter_rows(values_only=False):
        for cell in row:
            r = cell.row - 1
            c = cell.column - 1

            # Skip cells that are part of a merge but not the top-left
            if (cell.row, cell.column) in merge_skip:
                continue

            # Get/create format
            key = _cell_format_key(cell)
            if key != _DEFAULT_KEY:
                if key not in fmt_cache:
                    fmt_cache[key] = _make_format(workbook, key)
                fmt = fmt_cache[key]
            else:
                fmt = None

            # Set row height if customized
            row_dim = src_ws.row_dimensions.get(cell.row)
            if row_dim and row_dim.height and c == 0:
                opts = {"hidden": True} if row_dim.hidden else {}
                dst_ws.set_row(r, row_dim.height, None, opts)
            elif row_dim and row_dim.hidden and c == 0:
                dst_ws.set_row(r, None, None, {"hidden": True})

            # Handle merged cells
            is_merge_origin = False
            for (r1, c1, r2, c2) in merged_cells:
                if cell.row == r1 and cell.column == c1:
                    is_merge_origin = True
                    val = cell.value
                    if val is None:
                        val = ""
                    dst_ws.merge_range(r1 - 1, c1 - 1, r2 - 1, c2 - 1, val, fmt)
                    break

            if is_merge_origin:
                continue

            # Write regular cell
            if cell.value is not None:
                _write_cell(dst_ws, r, c, cell.value, fmt)
            elif fmt:
                dst_ws.write_blank(r, c, None, fmt)
