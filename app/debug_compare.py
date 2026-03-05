"""Compare our output vs target to find exactly which 2 rows differ in mapped status."""
import openpyxl

TARGET = "/Users/matthew.lucas/Downloads/Activity Aggregator - 03.06.2026 Forecast .xlsx"
OURS = "/Users/matthew.lucas/Documents/Activity Aggregator Update/app/output/"

import os, glob

# Find our latest output
outputs = glob.glob(os.path.join(OURS, "*.xlsx"))
if not outputs:
    print("No output file found. Run the pipeline first.")
    exit(1)
ours_path = max(outputs, key=os.path.getmtime)
print(f"Our output: {ours_path}")
print(f"Target:     {TARGET}")

def load_mapped_rows(path, label):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Alteryx_Output"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    col_idx = {h: i for i, h in enumerate(headers) if h}

    mapped = []
    not_mapped = []
    excluded = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        muc = row[col_idx.get("Manual User Check", -1)] if "Manual User Check" in col_idx else ""
        if muc == "Mapped":
            mapped.append(row)
        elif muc == "Not Mapped - Check":
            not_mapped.append(row)
        else:
            excluded.append(row)

    wb.close()
    print(f"\n{label}: Mapped={len(mapped)}, NotMapped={len(not_mapped)}, Excluded={len(excluded)}, Total={len(mapped)+len(not_mapped)+len(excluded)}")
    return headers, col_idx, mapped, not_mapped

# Load both
t_headers, t_idx, t_mapped, t_not_mapped = load_mapped_rows(TARGET, "TARGET")
o_headers, o_idx, o_mapped, o_not_mapped = load_mapped_rows(ours_path, "OURS")

# Build fingerprints for mapped rows
def fingerprint(row, idx):
    source = row[idx.get("Source", -1)] or ""
    acct = row[idx.get("Account Name", -1)] or ""
    asset = row[idx.get("Asset Code", -1)] or ""
    net = row[idx.get("Net Activity - USD", -1)] or 0
    mapping = row[idx.get("13WCF Line Item Mapping", -1)] or ""
    ref = row[idx.get("13WCF Ref #", -1)] or ""
    notes = row[idx.get("Notes", -1)] or ""
    return (source, acct, asset, net, mapping, ref, str(notes)[:50])

def fingerprint_key(row, idx):
    """Simpler key for set comparison."""
    source = row[idx.get("Source", -1)] or ""
    acct = row[idx.get("Account Name", -1)] or ""
    asset = row[idx.get("Asset Code", -1)] or ""
    net = row[idx.get("Net Activity - USD", -1)] or 0
    try:
        net = round(float(net), 2)
    except:
        pass
    notes = str(row[idx.get("Notes", -1)] or "")[:50]
    return (source, acct, asset, net, notes)

# Compare by Source + Mapping distribution
print("\n--- TARGET mapped by Source+Mapping ---")
from collections import Counter
t_dist = Counter()
for r in t_mapped:
    src = r[t_idx.get("Source", -1)] or ""
    mp = r[t_idx.get("13WCF Line Item Mapping", -1)] or ""
    t_dist[(src, mp)] += 1

o_dist = Counter()
for r in o_mapped:
    src = r[o_idx.get("Source", -1)] or ""
    mp = r[o_idx.get("13WCF Line Item Mapping", -1)] or ""
    o_dist[(src, mp)] += 1

all_keys = sorted(set(t_dist.keys()) | set(o_dist.keys()))
print(f"\n{'Source':<15} {'Mapping':<40} {'Target':>7} {'Ours':>7} {'Diff':>7}")
print("-" * 90)
for key in all_keys:
    tc = t_dist.get(key, 0)
    oc = o_dist.get(key, 0)
    diff = oc - tc
    marker = " ***" if diff != 0 else ""
    print(f"{key[0]:<15} {key[1]:<40} {tc:>7} {oc:>7} {diff:>+7}{marker}")

# Find rows in ours-mapped but not in target-mapped
print("\n\n--- ROWS MAPPED IN OURS BUT NOT IN TARGET ---")
t_keys = Counter()
for r in t_mapped:
    t_keys[fingerprint_key(r, t_idx)] += 1

o_keys = Counter()
for r in o_mapped:
    o_keys[fingerprint_key(r, o_idx)] += 1

extra_in_ours = o_keys - t_keys
if extra_in_ours:
    for fp, count in extra_in_ours.most_common():
        print(f"  [{count}x] Source={fp[0]}, Acct={fp[1]}, Asset={fp[2]}, Net={fp[3]}, Notes={fp[4]}")
else:
    print("  None found with simple fingerprint")

print("\n--- ROWS MAPPED IN TARGET BUT NOT IN OURS ---")
extra_in_target = t_keys - o_keys
if extra_in_target:
    for fp, count in extra_in_target.most_common():
        print(f"  [{count}x] Source={fp[0]}, Acct={fp[1]}, Asset={fp[2]}, Net={fp[3]}, Notes={fp[4]}")
else:
    print("  None found with simple fingerprint")
